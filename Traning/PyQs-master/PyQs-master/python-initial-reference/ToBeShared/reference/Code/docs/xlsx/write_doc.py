from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()
dest_filename = 'empty_book.xlsx'

ws1 = wb.active
ws1.title = "range names"

#Append complete row, 0-600 for 40 rows 
for row in range(1, 40):
    ws1.append(range(600))

ws2 = wb.create_sheet(title="Pi")
#write directly to F5 (columnRow)
ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="Data")
#Write specifically 
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))

print(ws3['AA10'].value)

#More writing 
ws = wb.create_sheet(title="Misc")
# set date using a Python datetime
import datetime 
ws['A1'] = datetime.datetime(2010, 7, 21)
ws['A1'].number_format #'yyyy-mm-dd h:mm:ss'
# add a simple formula
ws["A2"] = "=SUM(1, 1)" #or "=SUM(A1:B6)"

#openpyxl never evaluates formula but it is possible to check the name of a formula:
from openpyxl.utils import FORMULAE
"HEX2DEC" in FORMULAE #True 

#Merge / Unmerge cells
#When you merge cells all cells but the top-left one are removed 
ws.merge_cells('A3:D4')
ws.unmerge_cells('A3:D4')
# or equivalently, index starts from 1
ws.merge_cells(start_row=3, start_column=1, end_row=4, end_column=4)
#ws.unmerge_cells(start_row=2, start_column=1, end_row=4, end_column=4)

#Inserting an image
from openpyxl.drawing.image import Image
ws['A5'] = 'You should below'
img = Image('../data/monty-truth.png')
# add to worksheet and anchor next to cells
ws.add_image(img, 'A6')

#Line chart 
from datetime import date
from openpyxl.chart import (
    LineChart,
    Reference,
    BarChart,
    PieChart
)
from openpyxl.chart.axis import DateAxis

rows = [
    ['Date', 'Batch 1', 'Batch 2', 'Batch 3'],
    [date(2015,9, 1), 40, 30, 25],
    [date(2015,9, 2), 40, 25, 30],
    [date(2015,9, 3), 50, 30, 45],
    [date(2015,9, 4), 30, 25, 40],
    [date(2015,9, 5), 25, 35, 30],
    [date(2015,9, 6), 20, 40, 35],
]

rows  += [
    ['Pie', 'Sold'],
    ['Apple', 50],
    ['Cherry', 30],
    ['Pumpkin', 10],
    ['Chocolate', 40],
]

ws = wb.create_sheet(title="Charts")
for row in rows:
    ws.append(row) #starts from A1 to D7, other one A8 to B12

#Put some border 
from openpyxl.styles.borders import Border, Side
thin_border = Border(
    left=Side(border_style="thin", color='00000000'),
    right=Side(border_style="thin", color='00000000'),
    top=Side(border_style="thin", color='00000000'),
    bottom=Side(border_style="thin", color='00000000')
)
#from A1 to D7 
#ws.cell(row=3, column=2).border = thin_border
for row in ws['A1': 'B12']:
    for cell in row:
        cell.border = thin_border

for row in ws['C1': 'D7']:
    for cell in row:
        cell.border = thin_border

#Bold header , note style is immutable 
for row in ws['A1': 'D1']:
    for cell in row:
        cell.font = cell.font.copy(bold=True, italic=True)

for row in ws['A8': 'B8']:
    for cell in row:
        cell.font = cell.font.copy(bold=True)


#line chart 
c1 = LineChart()
c1.title = "Sales"
c1.y_axis.title = 'Size'

#crossAx is must (this from excel specification)
#Specifies the axId(axis ID) of axis that this axis cross. 
c1.y_axis.crossAx = 500  #this is for dateAxis
c1.x_axis = DateAxis(crossAx=100) #this is for Numeric Axis , for TextAxis, it is 10, for SeriesAxis, it is 1000
c1.x_axis.number_format = 'd-mmm'
c1.x_axis.majorTimeUnit = "days"
c1.x_axis.title = "Date"
#index starts from 1, data is B1 to D7 including titles
data = Reference(ws, min_col=2, min_row=1, max_col=4, max_row=7)
c1.add_data(data, titles_from_data=True)
#now add date , A2 to A7 
dates = Reference(ws, min_col=1, min_row=2, max_row=7)
c1.set_categories(dates)

#By default the top-left corner of a chart is anchored to cell A15 
#and the size is 15 x 7.5 cm (approximately 5 columns by 14 rows). 
#This can be changed by setting the anchor, width and height properties of the chart.
ws.add_chart(c1, "A15")

#Bar-chart 
#index starts from 1,A8 to B12
labels = Reference(ws, min_col=1, min_row=9, max_row=12)
#labels
#'Charts'!$A$9:$A$12
data = Reference(ws, min_col=2, min_row=8, max_row=12)

c2 = BarChart()
c2.add_data(data, titles_from_data=True)
c2.set_categories(labels)
c2.title = "Pies sold by category"
ws.add_chart(c2, "F15")

#Pie chart 
c3 = PieChart()
c3.add_data(data, titles_from_data=True)
c3.set_categories(labels)
c3.title = "Pies sold by category"
ws.add_chart(c3, "L15")

#now Save 
wb.save(filename = dest_filename)