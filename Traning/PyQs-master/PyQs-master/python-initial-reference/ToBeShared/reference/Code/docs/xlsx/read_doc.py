from openpyxl import load_workbook

#load_workbook options 
#guess_types will enable or disable (default) type inference when reading cells.
#data_only controls whether cells with formulae have either the formula (default) 
#or the value stored the last time Excel read the sheet.
#keep_vba controls whether any Visual Basic elements are preserved or not (default). 
#If they are preserved they are still not editable.



wb = load_workbook(filename = '../data/empty_book.xlsx')
sheet_ranges = wb['range names'] #sheet Name 

print("""
Workbook attributes 
""")
print([a for a in dir(wb) if a.startswith("_") == False])


print("accessing one value:")
print(sheet_ranges['D18'].value) #3

#Sheets properties 
print("sheets:")
print(wb.get_sheet_names())

active_sheet = wb.active
print("active_sheet type:")
print(type(active_sheet))
print("""
WorkSheet attributes 
""")
print([a for a in dir(active_sheet) if a.startswith("_") == False])

print("get one sheet:")
sheet = wb.get_sheet_by_name("Charts")
print(sheet.title)

#Range 
print("Read Range")
cells = sheet_ranges['A1': 'B6']
for c1, c2 in cells:
    print("{0:8} {1:8}".format(c1.value, c2.value))
    
print("other way access")
sheet['A1'] #<Cell 'Sheet'.A1>
sheet['A1'].value #note only cell has .value 

sheet['A1': 'D25'] #((<Cell 'Sheet'.A1>, <Cell 'Sheet'.B1>,...),...)
#or 
sheet['A1:D25']  #((<Cell 'Sheet'.A1>, <Cell 'Sheet'.B1>,...),...)
[ [c.value  for c in r]   for r in sheet['A1': 'D25']]
 
sheet['A']
sheet[4]  #index starts from 1 #(<Cell 'Sheet'.A4>, <Cell 'Sheet'.B4>...)
sheet['A:D']
sheet['A':'D']
sheet[1:4]

#Others are 
list(sheet.values)
list(sheet.rows)
list(sheet.columns)

print("Iterating by rows, ie row by row ")
#index from 1
for row in sheet_ranges.iter_rows(min_row=1, min_col=1, max_row=6, max_col=3):
    for cell in row:
        print(cell.value, end=" ")
    print()   

#0 1 2
#0 1 2
#0 1 2
print("Iterating by columns ie column by column ")
for col in sheet_ranges.iter_cols(min_row=1, min_col=1, max_row=6, max_col=3):
    for cell in col:
        print(cell.value, end=" ")
    print()    

#0 0 0 
#1 1 1 
#2 2 2 

print("simple stats ")
import statistics as stats
rows = sheet_ranges.rows

values = []
for row in rows:
    for cell in row:
        values.append(cell.value)

print("Number of values: {0}".format(len(values)))
print("Sum of values: {0}".format(sum(values)))
print("Minimum value: {0}".format(min(values)))
print("Maximum value: {0}".format(max(values)))
print("Mean: {0}".format(stats.mean(values)))
print("Median: {0}".format(stats.median(values)))
print("Standard deviation: {0}".format(stats.stdev(values)))
print("Variance: {0}".format(stats.variance(values)))

print("""
Dimensions
To get those cells that actually contain data, we can use dimensions.
""")

print(sheet.dimensions) #A1:WB39
print("Minimum row: {0}".format(sheet.min_row))
print("Maximum row: {0}".format(sheet.max_row))
print("Minimum column: {0}".format(sheet.min_column))
print("Maximum column: {0}".format(sheet.max_column))

values = []
for row in sheet[sheet.dimensions]:
    values.append([cell.value for cell in row])

print(values)


  